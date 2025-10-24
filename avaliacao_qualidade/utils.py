from django.http import HttpResponse
from django.template.loader import get_template
from django.db.models import Avg, Count
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, Reference
from io import BytesIO
import logging

logger = logging.getLogger(__name__)


def gerar_relatorio_pdf(avaliacao_config):
    """Gera relatório em PDF para uma avaliação"""
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center
        textColor=colors.HexColor('#1f4e79')
    )
    
    story.append(Paragraph("RELATÓRIO DE AVALIAÇÃO DE QUALIDADE", title_style))
    story.append(Paragraph("FATESA - Faculdade de Tecnologia em Saúde", styles['Heading2']))
    story.append(Spacer(1, 20))
    
    # Informações da avaliação
    info_data = [
        ['Curso:', avaliacao_config.curso.nome],
        ['Coordenador:', avaliacao_config.coordenador.nome],
        ['Turma:', avaliacao_config.turma],
        ['Data de Criação:', avaliacao_config.data_criacao.strftime('%d/%m/%Y')],
        ['Total de Respostas:', str(avaliacao_config.total_avaliacoes)]
    ]
    
    # Professores
    professores_nomes = ', '.join([p.nome for p in avaliacao_config.professores.all()])
    info_data.append(['Professores:', professores_nomes])
    
    info_table = Table(info_data, colWidths=[2*inch, 4*inch])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e6f2ff')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    
    story.append(info_table)
    story.append(Spacer(1, 20))
    
    # Estatísticas
    respostas = avaliacao_config.respostas.all()
    
    if respostas.exists():
        stats = respostas.aggregate(
            media_relacionamento=Avg('nota_relacionamento_professor'),
            media_didatica=Avg('nota_didatica_professor'),
            media_dominio=Avg('nota_dominio_assunto'),
            media_teorico=Avg('nota_conteudo_teorico'),
            media_pratico=Avg('nota_atividade_pratica'),
            media_portaria=Avg('nota_portaria'),
            media_atendimento=Avg('nota_atendimento_aluno'),
            media_secretaria=Avg('nota_secretaria'),
            media_recepcao=Avg('nota_recepcao_paciente'),
            media_biblioteca=Avg('nota_biblioteca'),
            media_comercial=Avg('nota_setor_comercial'),
            media_limpeza=Avg('nota_limpeza'),
            media_cantina=Avg('nota_cantina')
        )
        
        story.append(Paragraph("MÉDIAS POR CATEGORIA", styles['Heading2']))
        story.append(Spacer(1, 10))
        
        # Tabela de médias
        medias_data = [
            ['CATEGORIA', 'MÉDIA'],
            ['Relacionamento Professor-Aluno', f"{stats['media_relacionamento']:.1f}" if stats['media_relacionamento'] else "N/A"],
            ['Didática dos Professores', f"{stats['media_didatica']:.1f}" if stats['media_didatica'] else "N/A"],
            ['Domínio do Assunto', f"{stats['media_dominio']:.1f}" if stats['media_dominio'] else "N/A"],
            ['Conteúdo Teórico', f"{stats['media_teorico']:.1f}" if stats['media_teorico'] else "N/A"],
            ['Atividade Prática', f"{stats['media_pratico']:.1f}" if stats['media_pratico'] else "N/A"],
            ['Portaria', f"{stats['media_portaria']:.1f}" if stats['media_portaria'] else "N/A"],
            ['Atendimento ao Aluno', f"{stats['media_atendimento']:.1f}" if stats['media_atendimento'] else "N/A"],
            ['Secretaria', f"{stats['media_secretaria']:.1f}" if stats['media_secretaria'] else "N/A"],
            ['Recepção Paciente', f"{stats['media_recepcao']:.1f}" if stats['media_recepcao'] else "N/A"],
            ['Biblioteca', f"{stats['media_biblioteca']:.1f}" if stats['media_biblioteca'] else "N/A"],
            ['Setor Comercial', f"{stats['media_comercial']:.1f}" if stats['media_comercial'] else "N/A"],
            ['Limpeza', f"{stats['media_limpeza']:.1f}" if stats['media_limpeza'] else "N/A"],
            ['Cantina', f"{stats['media_cantina']:.1f}" if stats['media_cantina'] else "N/A"],
        ]
        
        medias_table = Table(medias_data, colWidths=[4*inch, 1*inch])
        medias_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(medias_table)
        story.append(Spacer(1, 20))
        
        # Comentários (se houver)
        comentarios = respostas.exclude(comentarios_adicionais__isnull=True).exclude(comentarios_adicionais='')
        if comentarios.exists():
            story.append(Paragraph("COMENTÁRIOS DOS ALUNOS", styles['Heading2']))
            story.append(Spacer(1, 10))
            
            for i, resposta in enumerate(comentarios[:10], 1):  # Máximo 10 comentários
                nome = resposta.nome_aluno or "Anônimo"
                comentario = resposta.comentarios_adicionais[:500]  # Limitar tamanho
                
                story.append(Paragraph(f"<b>{i}. {nome}:</b>", styles['Normal']))
                story.append(Paragraph(comentario, styles['Normal']))
                story.append(Spacer(1, 10))
    
    else:
        story.append(Paragraph("Nenhuma resposta encontrada para esta avaliação.", styles['Normal']))
    
    # Rodapé
    story.append(Spacer(1, 30))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        alignment=1,
        textColor=colors.grey
    )
    story.append(Paragraph(f"Relatório gerado em {avaliacao_config.data_criacao.strftime('%d/%m/%Y %H:%M')}", footer_style))
    
    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content


def gerar_relatorio_excel(avaliacao_config):
    """Gera relatório em Excel para uma avaliação"""
    
    wb = openpyxl.Workbook()
    
    # === ABA 1: INFORMAÇÕES GERAIS ===
    ws_info = wb.active
    ws_info.title = "Informações Gerais"
    
    # Cabeçalho
    ws_info['A1'] = "RELATÓRIO DE AVALIAÇÃO DE QUALIDADE - FATESA"
    ws_info['A1'].font = Font(size=16, bold=True, color="1F4E79")
    ws_info.merge_cells('A1:D1')
    
    # Informações da avaliação
    info_data = [
        ['Curso:', avaliacao_config.curso.nome],
        ['Coordenador:', avaliacao_config.coordenador.nome],
        ['Turma:', avaliacao_config.turma],
        ['Data de Criação:', avaliacao_config.data_criacao.strftime('%d/%m/%Y')],
        ['Total de Respostas:', avaliacao_config.total_avaliacoes]
    ]
    
    # Professores
    professores_nomes = ', '.join([p.nome for p in avaliacao_config.professores.all()])
    info_data.append(['Professores:', professores_nomes])
    
    row = 3
    for label, value in info_data:
        ws_info[f'A{row}'] = label
        ws_info[f'A{row}'].font = Font(bold=True)
        ws_info[f'B{row}'] = value
        row += 1
    
    # Ajustar largura das colunas
    ws_info.column_dimensions['A'].width = 20
    ws_info.column_dimensions['B'].width = 50
    
    # === ABA 2: ESTATÍSTICAS ===
    ws_stats = wb.create_sheet("Estatísticas")
    
    respostas = avaliacao_config.respostas.all()
    
    if respostas.exists():
        stats = respostas.aggregate(
            media_relacionamento=Avg('nota_relacionamento_professor'),
            media_didatica=Avg('nota_didatica_professor'),
            media_dominio=Avg('nota_dominio_assunto'),
            media_teorico=Avg('nota_conteudo_teorico'),
            media_pratico=Avg('nota_atividade_pratica'),
            media_portaria=Avg('nota_portaria'),
            media_atendimento=Avg('nota_atendimento_aluno'),
            media_secretaria=Avg('nota_secretaria'),
            media_recepcao=Avg('nota_recepcao_paciente'),
            media_biblioteca=Avg('nota_biblioteca'),
            media_comercial=Avg('nota_setor_comercial'),
            media_limpeza=Avg('nota_limpeza'),
            media_cantina=Avg('nota_cantina')
        )
        
        # Cabeçalhos
        ws_stats['A1'] = "CATEGORIA"
        ws_stats['B1'] = "MÉDIA"
        ws_stats['A1'].font = Font(bold=True)
        ws_stats['B1'].font = Font(bold=True)
        
        # Dados
        categorias_medias = [
            ('Relacionamento Professor-Aluno', stats['media_relacionamento']),
            ('Didática dos Professores', stats['media_didatica']),
            ('Domínio do Assunto', stats['media_dominio']),
            ('Conteúdo Teórico', stats['media_teorico']),
            ('Atividade Prática', stats['media_pratico']),
            ('Portaria', stats['media_portaria']),
            ('Atendimento ao Aluno', stats['media_atendimento']),
            ('Secretaria', stats['media_secretaria']),
            ('Recepção Paciente', stats['media_recepcao']),
            ('Biblioteca', stats['media_biblioteca']),
            ('Setor Comercial', stats['media_comercial']),
            ('Limpeza', stats['media_limpeza']),
            ('Cantina', stats['media_cantina']),
        ]
        
        row = 2
        for categoria, media in categorias_medias:
            ws_stats[f'A{row}'] = categoria
            ws_stats[f'B{row}'] = round(media, 1) if media else "N/A"
            row += 1
        
        # Formatação
        ws_stats.column_dimensions['A'].width = 30
        ws_stats.column_dimensions['B'].width = 15
        
        # Gráfico (opcional - requer mais configuração)
        # chart = BarChart()
        # chart.title = "Médias por Categoria"
        # data = Reference(ws_stats, min_col=2, min_row=1, max_row=row-1)
        # categories = Reference(ws_stats, min_col=1, min_row=2, max_row=row-1)
        # chart.add_data(data, titles_from_data=True)
        # chart.set_categories(categories)
        # ws_stats.add_chart(chart, "D2")
    
    # === ABA 3: RESPOSTAS DETALHADAS ===
    ws_respostas = wb.create_sheet("Respostas Detalhadas")
    
    if respostas.exists():
        # Cabeçalhos
        headers = [
            'Data', 'Nome', 'Contato', 'Relacionamento Prof.', 'Didática Prof.',
            'Domínio Assunto', 'Respeita Horários', 'Origem', 'Motivo Escolha',
            'Conteúdo Teórico', 'Atividade Prática', 'Portaria', 'Atendimento',
            'Secretaria', 'Recepção', 'Biblioteca', 'Comercial', 'Limpeza',
            'Cantina', 'Comentários', 'Sugestões'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws_respostas.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid")
        
        # Dados das respostas
        for row, resposta in enumerate(respostas, 2):
            data = [
                resposta.data_resposta.strftime('%d/%m/%Y %H:%M'),
                resposta.nome_aluno or 'Anônimo',
                resposta.contato_aluno or '',
                resposta.nota_relacionamento_professor,
                resposta.nota_didatica_professor,
                resposta.nota_dominio_assunto,
                'Sim' if resposta.professor_respeita_horarios else 'Não',
                resposta.get_origem_conhecimento_display(),
                resposta.get_motivo_escolha_display(),
                resposta.nota_conteudo_teorico,
                resposta.nota_atividade_pratica,
                resposta.nota_portaria,
                resposta.nota_atendimento_aluno,
                resposta.nota_secretaria,
                resposta.nota_recepcao_paciente,
                resposta.nota_biblioteca,
                resposta.nota_setor_comercial,
                resposta.nota_limpeza,
                resposta.nota_cantina,
                resposta.comentarios_adicionais or '',
                resposta.sugestoes_melhorias or ''
            ]
            
            for col, value in enumerate(data, 1):
                ws_respostas.cell(row=row, column=col, value=value)
        
        # Ajustar largura das colunas
        for col in range(1, len(headers) + 1):
            ws_respostas.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Salvar em buffer
    buffer = BytesIO()
    wb.save(buffer)
    excel_content = buffer.getvalue()
    buffer.close()
    
    return excel_content


def calcular_estatisticas_periodo(data_inicio, data_fim, filtros=None):
    """Calcula estatísticas para um período específico"""
    
    from .models import AvaliacaoResposta
    
    respostas = AvaliacaoResposta.objects.filter(
        data_resposta__date__gte=data_inicio,
        data_resposta__date__lte=data_fim
    )
    
    # Aplicar filtros adicionais
    if filtros:
        if filtros.get('curso'):
            respostas = respostas.filter(avaliacao_config__curso=filtros['curso'])
        if filtros.get('coordenador'):
            respostas = respostas.filter(avaliacao_config__coordenador=filtros['coordenador'])
        if filtros.get('professor'):
            respostas = respostas.filter(avaliacao_config__professores=filtros['professor'])
    
    if not respostas.exists():
        return None
    
    stats = respostas.aggregate(
        total_respostas=Count('id'),
        media_professor=Avg('nota_relacionamento_professor'),
        media_didatica=Avg('nota_didatica_professor'),
        media_dominio=Avg('nota_dominio_assunto'),
        media_curso_teorico=Avg('nota_conteudo_teorico'),
        media_curso_pratico=Avg('nota_atividade_pratica'),
        media_administracao=Avg('nota_portaria')
    )
    
    # Calcular média geral
    medias_individuais = [
        stats['media_professor'],
        stats['media_didatica'],
        stats['media_dominio'],
        stats['media_curso_teorico'],
        stats['media_curso_pratico'],
        stats['media_administracao']
    ]
    
    medias_validas = [m for m in medias_individuais if m is not None]
    stats['media_geral'] = sum(medias_validas) / len(medias_validas) if medias_validas else 0
    
    return stats


def gerar_relatorio_comparativo(avaliacoes_configs):
    """Gera relatório comparativo entre múltiplas avaliações"""
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,
        textColor=colors.HexColor('#1f4e79')
    )
    
    story.append(Paragraph("RELATÓRIO COMPARATIVO DE AVALIAÇÕES", title_style))
    story.append(Paragraph("FATESA - Faculdade de Tecnologia em Saúde", styles['Heading2']))
    story.append(Spacer(1, 20))
    
    # Tabela comparativa
    data = [['Curso', 'Turma', 'Respostas', 'Média Geral']]
    
    for config in avaliacoes_configs:
        respostas = config.respostas.all()
        if respostas.exists():
            media_geral = sum([
                r.get_media_geral() for r in respostas
            ]) / respostas.count()
        else:
            media_geral = 0
        
        data.append([
            config.curso.nome[:30] + '...' if len(config.curso.nome) > 30 else config.curso.nome,
            config.turma,
            str(config.total_avaliacoes),
            f"{media_geral:.1f}"
        ])
    
    table = Table(data, colWidths=[3*inch, 1.5*inch, 1*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4e79')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    story.append(table)
    
    doc.build(story)
    pdf_content = buffer.getvalue()
    buffer.close()
    
    return pdf_content